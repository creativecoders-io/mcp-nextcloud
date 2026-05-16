#!/usr/bin/env python3
"""
Nextcloud MCP Server

Provides tools for creating, reading, and managing files in Nextcloud via WebDAV.
"""

import os
import sys
import json
import logging
from typing import Any
from io import BytesIO

from webdav3.client import Client
from openpyxl import Workbook
from openpyxl.styles import Font
from mcp.server import Server
from mcp.types import Tool, TextContent

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configuration from environment variables
NEXTCLOUD_URL = os.getenv('NEXTCLOUD_URL')
NEXTCLOUD_USER = os.getenv('NEXTCLOUD_USER', 'assistant')
NEXTCLOUD_PASSWORD = os.getenv('NEXTCLOUD_APP_PASSWORD')

if not NEXTCLOUD_URL:
    logger.error("NEXTCLOUD_URL environment variable is required")
    sys.exit(1)

if not NEXTCLOUD_PASSWORD:
    logger.error("NEXTCLOUD_APP_PASSWORD environment variable is required")
    sys.exit(1)

# WebDAV configuration
webdav_options = {
    'webdav_hostname': f"{NEXTCLOUD_URL}/remote.php/dav/files/{NEXTCLOUD_USER}",
    'webdav_login': NEXTCLOUD_USER,
    'webdav_password': NEXTCLOUD_PASSWORD,
    'disable_check': True  # Skip SSL verification if needed (adjust for production)
}

client = Client(webdav_options)


def list_folder(path: str) -> list[dict[str, Any]]:
    """List files and folders in a directory."""
    logger.info(f"Listing folder: {path}")
    
    try:
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        
        items = client.list(path, get_info=True)
        
        result = []
        for item in items:
            if item == path.rstrip('/') + '/':
                continue  # Skip the folder itself
            
            info = client.info(item)
            result.append({
                'name': os.path.basename(item.rstrip('/')),
                'path': item,
                'type': 'directory' if info.get('isdir') else 'file',
                'size': info.get('size', 0),
                'modified': info.get('modified', '')
            })
        
        return result
    
    except Exception as e:
        logger.error(f"Error listing folder: {e}")
        raise


def read_file(path: str) -> dict[str, Any]:
    """Read file contents."""
    logger.info(f"Reading file: {path}")
    
    try:
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        
        content = client.resource(path).read()
        
        # Try to decode as text
        try:
            content_str = content.decode('utf-8')
        except UnicodeDecodeError:
            content_str = f"<binary file, {len(content)} bytes>"
        
        info = client.info(path)
        
        return {
            'content': content_str,
            'path': path,
            'size': info.get('size', len(content)),
            'modified': info.get('modified', '')
        }
    
    except Exception as e:
        logger.error(f"Error reading file: {e}")
        raise


def create_file(path: str, content: str, mime_type: str | None = None) -> dict[str, Any]:
    """Create or update a file."""
    logger.info(f"Creating file: {path}")
    
    try:
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        
        # Ensure parent directory exists
        parent_dir = os.path.dirname(path)
        if parent_dir and not client.check(parent_dir):
            create_folder(parent_dir)
        
        # Write content
        content_bytes = content.encode('utf-8')
        client.resource(path).write(content_bytes)
        
        return {
            'path': path,
            'url': f"{NEXTCLOUD_URL}/f/{path}",
            'size': len(content_bytes)
        }
    
    except Exception as e:
        logger.error(f"Error creating file: {e}")
        raise


def create_excel(path: str, data: dict[str, Any]) -> dict[str, Any]:
    """Create an Excel file with structured data."""
    logger.info(f"Creating Excel file: {path}")
    
    try:
        # Ensure path starts with / and ends with .xlsx
        if not path.startswith('/'):
            path = '/' + path
        if not path.endswith('.xlsx'):
            path += '.xlsx'
        
        # Ensure parent directory exists
        parent_dir = os.path.dirname(path)
        if parent_dir and not client.check(parent_dir):
            create_folder(parent_dir)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        total_rows = 0
        
        # Add sheets
        for sheet_data in data.get('sheets', []):
            ws = wb.create_sheet(title=sheet_data.get('name', 'Sheet1'))
            
            # Add header row
            columns = sheet_data.get('columns', [])
            ws.append(columns)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # Add data rows
            rows = sheet_data.get('rows', [])
            for row in rows:
                ws.append(row)
                total_rows += 1
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Upload to Nextcloud
        client.resource(path).write(buffer.read())
        
        return {
            'path': path,
            'url': f"{NEXTCLOUD_URL}/f/{path}",
            'sheets': len(data.get('sheets', [])),
            'rows': total_rows
        }
    
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise


def create_folder(path: str) -> dict[str, Any]:
    """Create a folder (and parent folders if needed)."""
    logger.info(f"Creating folder: {path}")
    
    try:
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        
        # Create folder (recursive)
        if not client.check(path):
            # Split path and create each level
            parts = [p for p in path.split('/') if p]
            current_path = ''
            
            for part in parts:
                current_path += '/' + part
                if not client.check(current_path):
                    client.mkdir(current_path)
        
        return {
            'path': path,
            'success': True
        }
    
    except Exception as e:
        logger.error(f"Error creating folder: {e}")
        raise


def search_files(query: str, path: str | None = None) -> list[dict[str, Any]]:
    """Search for files (simple implementation - lists and filters)."""
    logger.info(f"Searching for: {query} in path: {path or '/'}")
    
    try:
        search_path = path or '/'
        if not search_path.startswith('/'):
            search_path = '/' + search_path
        
        # Get all items recursively
        all_items = []
        
        def walk_dir(dir_path: str):
            try:
                items = client.list(dir_path, get_info=True)
                for item in items:
                    if item == dir_path.rstrip('/') + '/':
                        continue
                    
                    info = client.info(item)
                    if info.get('isdir'):
                        # Recursively search subdirectories
                        walk_dir(item)
                    else:
                        # Check if filename matches query
                        if query.lower() in os.path.basename(item).lower():
                            all_items.append({
                                'name': os.path.basename(item.rstrip('/')),
                                'path': item,
                                'type': 'file',
                                'size': info.get('size', 0),
                                'modified': info.get('modified', '')
                            })
            except Exception as e:
                logger.warning(f"Could not search directory {dir_path}: {e}")
        
        walk_dir(search_path)
        return all_items
    
    except Exception as e:
        logger.error(f"Error searching files: {e}")
        raise


# Create MCP server
app = Server("nextcloud-mcp")


@app.list_tools()
async def list_tools() -> list[Tool]:
    """List available tools."""
    return [
        Tool(
            name="nextcloud_list",
            description="List files and folders in a directory",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Folder path (e.g., '/Wijkfeest/documents')"
                    }
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="nextcloud_read_file",
            description="Read file contents",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "File path"
                    }
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="nextcloud_create_file",
            description="Create or update a file",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "File path"
                    },
                    "content": {
                        "type": "string",
                        "description": "File content"
                    },
                    "mime_type": {
                        "type": "string",
                        "description": "MIME type (optional, auto-detected from extension)"
                    }
                },
                "required": ["path", "content"]
            }
        ),
        Tool(
            name="nextcloud_create_excel",
            description="Create an Excel file with structured data",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "File path (will add .xlsx if not present)"
                    },
                    "data": {
                        "type": "object",
                        "description": "Excel data structure",
                        "properties": {
                            "sheets": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "name": {"type": "string"},
                                        "columns": {
                                            "type": "array",
                                            "items": {"type": "string"}
                                        },
                                        "rows": {
                                            "type": "array",
                                            "items": {"type": "array"}
                                        }
                                    }
                                }
                            }
                        }
                    }
                },
                "required": ["path", "data"]
            }
        ),
        Tool(
            name="nextcloud_create_folder",
            description="Create a folder (and parent folders if needed)",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Folder path"
                    }
                },
                "required": ["path"]
            }
        ),
        Tool(
            name="nextcloud_search",
            description="Search for files by name",
            inputSchema={
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search query (filename)"
                    },
                    "path": {
                        "type": "string",
                        "description": "Limit search to specific folder (optional)"
                    }
                },
                "required": ["query"]
            }
        )
    ]


@app.call_tool()
async def call_tool(name: str, arguments: Any) -> list[TextContent]:
    """Handle tool calls."""
    try:
        if name == "nextcloud_list":
            result = list_folder(path=arguments["path"])
        elif name == "nextcloud_read_file":
            result = read_file(path=arguments["path"])
        elif name == "nextcloud_create_file":
            result = create_file(
                path=arguments["path"],
                content=arguments["content"],
                mime_type=arguments.get("mime_type")
            )
        elif name == "nextcloud_create_excel":
            result = create_excel(
                path=arguments["path"],
                data=arguments["data"]
            )
        elif name == "nextcloud_create_folder":
            result = create_folder(path=arguments["path"])
        elif name == "nextcloud_search":
            result = search_files(
                query=arguments["query"],
                path=arguments.get("path")
            )
        else:
            raise ValueError(f"Unknown tool: {name}")
        
        return [TextContent(
            type="text",
            text=json.dumps(result, indent=2)
        )]
    
    except Exception as e:
        logger.error(f"Error executing tool {name}: {e}")
        return [TextContent(
            type="text",
            text=json.dumps({"error": str(e)}, indent=2)
        )]


if __name__ == "__main__":
    transport = os.getenv("MCP_TRANSPORT", "stdio")
    
    if transport == "http":
        logger.info("Starting Nextcloud MCP Server with Streamable HTTP transport...")
        app.run(
            transport="streamable-http",
            host="0.0.0.0",
            port=int(os.getenv("PORT", "8000")),
        )
    else:
        logger.info("Starting Nextcloud MCP Server with stdio transport...")
        app.run()
